#!/usr/bin/env python3
"""
pdf/pdf_generator.py
Domain-specific PDF generator for MDII evaluation reports
"""

import re
import time
from pathlib import Path
from typing import Dict, List, Optional
from collections import defaultdict

import openpyxl
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.colors import HexColor
from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_JUSTIFY
from reportlab.pdfgen import canvas

# PDF Configuration
PDF_CONFIG = {
    "early": {
        "DOMAIN_START_ROW": 16,
        "DOMAIN_COLUMNS": [1, 2, 3, 4, 5, 6],  # Columns A-F
        "FLAG_VALUE": "X"
    },
    "advanced": {
        "DOMAIN_START_ROW": 16,
        "DOMAIN_COLUMNS": [1, 2, 3, 4, 5, 6, 7],  # Columns A-G
        "FLAG_VALUE": "X"
    }
}

# Domain-specific subheadings that need numbering
DOMAIN_SUBHEADINGS = [
    "Engagement in Problem Definition", "Loss of Agency", "Content and Design",
    "Equality and Empowerment", "Diversity and Representation",
    "Community-led Solutions", "Infrastructure Readiness", "Integration with Existing Systems",
    "Resilience and Security", "Data Privacy", "Ethical Standards Adherence",
    "Ethical Oversight", "Impact Assessment", "Algorithmic Fairness",
    "Data Representation Equity", "Fraudulent activities", "Inability to Access Collected Data",
    "Unauthorized Access", "Bias Monitoring and Adaptation", "Long-term Viability",
    "Maintainability", "Scalability", "Problem-Solution Fit", "Adaptive Capability",
    "Problem Identification Accuracy", "Local Contextual Understanding", "Affordability"
]

def debug(msg: str):
    """Print debug messages"""
    print(msg)

def clean_html_text(html_text):
    """Clean HTML tags from text and preserve formatting"""
    if not html_text:
        return ""
    
    text = str(html_text)
    text = re.sub(r'<[^>]+>', ' ', text)
    text = re.sub(r'\s+', ' ', text)
    text = text.replace('&nbsp;', ' ')
    text = text.replace('&amp;', '&')
    text = text.replace('&lt;', '<')
    text = text.replace('&gt;', '>')
    
    return text.strip()

def analyze_domain_responses(ws_answers, maturity_key: str) -> Dict:
    """Analyze which domains have answered which questions"""
    debug("=== ANALYZING DOMAIN RESPONSES ===")
    
    config = PDF_CONFIG[maturity_key]
    domain_start_row = config["DOMAIN_START_ROW"]
    domain_columns = config["DOMAIN_COLUMNS"]
    flag_value = config["FLAG_VALUE"]
    
    # Get domain names from row 1
    domain_names = []
    for col_idx in domain_columns:
        domain_name = ws_answers.cell(1, col_idx).value
        if domain_name and str(domain_name).strip():
            domain_names.append((col_idx, str(domain_name).strip()))
            debug(f"Found domain in column {col_idx}: {domain_name}")
    
    if not domain_names:
        debug("No domain names found in header row")
        return {}
    
    # Analyze responses for each domain
    domain_responses = {}
    max_row = ws_answers.max_row or 1000
    
    for col_idx, domain_name in domain_names:
        answered_questions = []
        
        for row_idx in range(domain_start_row, max_row + 1):
            cell_value = ws_answers.cell(row_idx, col_idx).value
            
            if isinstance(cell_value, str) and cell_value.strip().upper() == flag_value:
                answered_questions.append(row_idx)
        
        if answered_questions:
            domain_responses[domain_name] = {
                'column': col_idx,
                'answered_rows': answered_questions
            }
            debug(f"Domain {domain_name} answered {len(answered_questions)} questions")
    
    return domain_responses

def get_html_questions_mapping(ws_html):
    """Create mapping of question codes to HTML formatted questions"""
    if not ws_html:
        return {}
    
    html_mapping = {}
    max_row = ws_html.max_row or 1000
    
    for row_idx in range(1, max_row + 1):
        for col_idx in range(16, 26):
            cell_value = ws_html.cell(row_idx, col_idx).value
            
            if cell_value and isinstance(cell_value, str):
                cell_text = str(cell_value).strip()
                code_matches = re.findall(r'\b(\d{8})\b', cell_text)
                
                if code_matches and ('html' in cell_text.lower() or '<' in cell_text):
                    for code in code_matches:
                        if code not in html_mapping:
                            html_mapping[code] = clean_html_text(cell_text)
    
    debug(f"Created HTML mapping for {len(html_mapping)} questions")
    return html_mapping

def identify_content_type(question_text: str, answer_text: str, previous_content_type: Optional[str] = None) -> str:
    """Enhanced content type identification"""
    if not question_text:
        return "question"
    
    clean_text = question_text.strip()
    clean_text_lower = clean_text.lower()
    
    # Rule 1: Text after Evaluator Affirmation is normal text
    if previous_content_type == "evaluator_affirmation":
        return "evaluator_affirmation_text"
    
    # Check for exact matches
    if clean_text == "Domain-Specific Questions":
        return "domain_heading"
    
    if clean_text == "ADDITIONAL INFORMATION":
        return "additional_info_heading"
    
    # Check for domain subheadings
    for subheading in DOMAIN_SUBHEADINGS:
        if subheading.strip() in clean_text:
            return "domain_subheading"
    
    # Keyword-based identification
    if "dimension:" in clean_text_lower and "sub-dimension" not in clean_text_lower:
        return "dimension"
    elif "sub-dimension" in clean_text_lower:
        return "subdimension"
    elif clean_text_lower.startswith("indicator:") or "indicator:" in clean_text_lower:
        return "indicator"
    elif clean_text_lower == "evaluator affirmation":
        return "evaluator_affirmation"
    elif clean_text_lower == "innovator response":
        return "innovator_response"
    elif "contextual information" in clean_text_lower:
        return "contextual_information"
    elif clean_text_lower.startswith("description"):
        return "description"
    else:
        return "question"

class NumberedCanvas(canvas.Canvas):
    """Custom canvas class for page numbers"""
    def __init__(self, *args, **kwargs):
        canvas.Canvas.__init__(self, *args, **kwargs)
        self._saved_page_states = []

    def showPage(self):
        self._saved_page_states.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        num_pages = len(self._saved_page_states)
        for page_num, state in enumerate(self._saved_page_states):
            self.__dict__.update(state)
            self.draw_page_number(page_num + 1, num_pages)
            canvas.Canvas.showPage(self)
        canvas.Canvas.save(self)

    def draw_page_number(self, page_num, total_pages):
        self.setFont("Helvetica", 9)
        self.setFillColor(HexColor('#6b7280'))
        self.drawRightString(A4[0] - inch*0.6, inch*0.4, f"{page_num}")

def create_pdf_styles():
    """Create all PDF styles"""
    styles = getSampleStyleSheet()
    
    custom_styles = {
        'title': ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=22,
            spaceAfter=15,
            spaceBefore=20,
            textColor=HexColor('#591fd5'),
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        ),
        'subtitle': ParagraphStyle(
            'CustomSubtitle',
            parent=styles['Normal'],
            fontSize=16,
            spaceAfter=25,
            spaceBefore=5,
            textColor=HexColor('#374151'),
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        ),
        'header': ParagraphStyle(
            'CustomHeader',
            parent=styles['Normal'],
            fontSize=10,
            spaceAfter=20,
            spaceBefore=10,
            alignment=TA_LEFT,
            fontName='Helvetica',
            bottomMargin=5,
        ),
        'thank_you': ParagraphStyle(
            'ThankYouNote',
            parent=styles['Normal'],
            fontSize=10,
            spaceBefore=15,
            alignment=TA_JUSTIFY,
            fontName='Helvetica',
        ),
        'section_header': ParagraphStyle(
            'SectionHeader',
            parent=styles['Heading2'],
            fontSize=16,
            spaceAfter=5,
            textColor=HexColor('#2b6cb0'),
            fontName='Helvetica-Bold',
        ),
        'section_desc': ParagraphStyle(
            'SectionDescription',
            parent=styles['Normal'],
            fontSize=10,
            alignment=TA_JUSTIFY,
            fontName='Helvetica',
        ),
        'main_heading': ParagraphStyle(
            'MainHeading',
            parent=styles['Heading2'],
            fontSize=14,
            spaceAfter=0,
            spaceBefore=10,
            textColor=HexColor('#1e3a8a'),
            fontName='Helvetica-Bold',
        ),
        'sub_heading': ParagraphStyle(
            'SubHeading',
            parent=styles['Heading3'],
            fontSize=13,
            spaceAfter=0,
            textColor=HexColor('#1e3a8a'),
            fontName='Helvetica-Bold',
        ),
        'indicator': ParagraphStyle(
            'IndicatorHeading',
            parent=styles['Heading4'],
            fontSize=12,
            spaceBefore=10,
            textColor=HexColor('#000000'),
            fontName='Helvetica-Bold',
        ),
        'evaluator': ParagraphStyle(
            'EvaluatorAffirmation',
            parent=styles['Heading4'],
            fontSize=11,
            textColor=HexColor('#1d4ed8'),
            fontName='Helvetica-Bold',
        ),
        'evaluator_text': ParagraphStyle(
            'EvaluatorText',
            parent=styles['Normal'],
            fontSize=10,
            spaceAfter=2,
            spaceBefore=0,
            textColor=HexColor('#000000'),
            fontName='Helvetica',
            alignment=TA_JUSTIFY
        ),
        'innovator': ParagraphStyle(
            'InnovatorResponse',
            parent=styles['Heading4'],
            fontSize=11,
            fontName='Helvetica-Bold',
        ),
        'context_heading': ParagraphStyle(
            'ContextHeading',
            parent=styles['Normal'],
            fontSize=11,
            spaceAfter=5,
            spaceBefore=12,
            textColor=HexColor('#374151'),
            fontName='Helvetica-Bold',
        ),
        'description': ParagraphStyle(
            'Description',
            parent=styles['Normal'],
            fontSize=10,
            alignment=TA_JUSTIFY,
            fontName='Helvetica',
            textColor=HexColor('#4b5563')
        ),
        'qa': ParagraphStyle(
            'QuestionAnswer',
            parent=styles['Normal'],
            fontSize=10,
            spaceAfter=5,
            spaceBefore=5,
            alignment=TA_JUSTIFY,
            fontName='Helvetica'
        ),
        'domain_heading': ParagraphStyle(
            'DomainHeading',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=15,
            spaceBefore=20,
            textColor=HexColor('#1e3a8a'),
            fontName='Helvetica-Bold',
        ),
        'domain_subheading': ParagraphStyle(
            'DomainSubheading',
            parent=styles['Heading3'],
            fontSize=14,
            spaceAfter=8,
            spaceBefore=12,
            textColor=HexColor('#2b6cb0'),
            fontName='Helvetica-Bold',
        ),
        'additional_info': ParagraphStyle(
            'AdditionalInfoHeading',
            parent=styles['Heading2'],
            fontSize=16,
            spaceAfter=10,
            spaceBefore=15,
            textColor=HexColor('#1e3a8a'),
            fontName='Helvetica-Bold',
        ),
        'footer': ParagraphStyle(
            'Footer',
            parent=styles['Normal'],
            fontSize=9,
            alignment=TA_CENTER,
            spaceAfter=10,
            textColor=HexColor('#6b7280')
        )
    }
    
    return custom_styles

def generate_pdfs_from_excel(tool_code: str, excel_path: Path, maturity_key: str) -> bool:
    """Generate domain-specific PDFs from Excel file"""
    debug(f"=== GENERATING DOMAIN PDFs ===")
    debug(f"Excel: {excel_path}")
    debug(f"Maturity: {maturity_key}")
    
    if not excel_path.exists():
        debug(f"Excel file not found: {excel_path}")
        return False
    
    # Setup output directory
    safe_tool_id = tool_code.replace("/", "_").replace("\\", "_").replace(":", "_")
    pdf_dir = excel_path.parent / "PDF"
    pdf_dir.mkdir(parents=True, exist_ok=True)
    
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        
        # Find sheets
        ws_answers = None
        for sheet_name in ["Innovator Answers", "InnovatorAnswers", "Innovator_Answers"]:
            if sheet_name in wb.sheetnames:
                ws_answers = wb[sheet_name]
                break
        
        if not ws_answers:
            ws_answers = wb.active
        
        ws_html = None
        for sheet_name in ["HTML Evaluation Compilation", "HTMLEvaluationCompilation", "UserTypeIV_Answers"]:
            if sheet_name in wb.sheetnames:
                ws_html = wb[sheet_name]
                break
        
        # Analyze domains
        domain_responses = analyze_domain_responses(ws_answers, maturity_key)
        
        if not domain_responses:
            debug("No domain responses found")
            wb.close()
            return False
        
        # Get HTML mapping
        html_mapping = get_html_questions_mapping(ws_html)
        
        # Get MDII version
        mdii_version = "Unknown"
        try:
            for sheet in wb.worksheets:
                for row_idx in range(1, 20):
                    for col_idx in range(1, 15):
                        cell_value = sheet.cell(row_idx, col_idx).value
                        if cell_value and isinstance(cell_value, str):
                            if "mdii version" in str(cell_value).lower():
                                mdii_version = str(cell_value).strip()
                                break
        except:
            pass
        
        # Create styles
        styles = create_pdf_styles()
        
        pdf_count = 0
        
        # Generate PDF for each domain
        for domain_name, domain_info in domain_responses.items():
            try:
                debug(f"Creating PDF for domain: {domain_name}")
                
                pdf_name = f"{safe_tool_id}_{domain_name.replace(' ', '_')}_MDII_Evaluation_Report.pdf"
                pdf_path = pdf_dir / pdf_name
                
                # Create PDF document
                doc = SimpleDocTemplate(
                    str(pdf_path),
                    pagesize=A4,
                    rightMargin=inch*0.6,
                    leftMargin=inch*0.6,
                    topMargin=inch*0.8,
                    bottomMargin=inch*0.8
                )
                
                story = []
                
                # Title page
                story.append(Paragraph("Multidimensional Digital Inclusiveness Assessment", styles['title']))
                story.append(Paragraph("Domain Evaluation Report", styles['subtitle']))
                story.append(Spacer(1, 20))
                
                # Header information
                header_info = f"""
                <b><font color='#2b6cb0' size='10'>Digital Tool Code:</font></b> <font color='#1a202c'>{tool_code}</font><br/>
                <b><font color='#2b6cb0' size='10'>MDII Version:</font></b> <font color='#1a202c'>{mdii_version}</font><br/>
                <b><font color='#2b6cb0' size='10'>Domain Expert:</font></b> <font color='#1a202c'>{domain_name}</font><br/>
                <b><font color='#2b6cb0' size='10'>Report Generated:</font></b> <font color='#1a202c'>{time.strftime('%B %d, %Y at %H:%M')}</font><br/>
                """
                story.append(Paragraph(header_info, styles['header']))
                
                # Thank you note
                thank_you_text = """
                <b>Thank you for your time</b><br/><br/>
                You are receiving this form because you have been identified as a relevant expert for the evaluation of digital inclusiveness of this digital tool;<br/><br/>
                Your answers will help the calculation of the Multidimensional Digital Inclusiveness Index (MDII). The MDII is a tool designed to scientifically evaluate and enhance the digital inclusiveness of agritools for marginalized groups, across seven dimensions: Beneficial Impact, Risks & Harms, Accessibility, Usage Effectiveness, Supportive Ecosystem, Ethical and Responsible Innovation, and Co-creation and Governance.<br/><br/>
                This document is a compilation of the answers provided by the Innovator of the tool and divided in 2 sections: (I) General; (II) Domain-Specific. The affirmations for evaluation are in blue. Next to it, you'll find the information that was provided, as well as the innovators' answers. No changes were made on the answers — this means what you'll be reading is raw data.
                """
                story.append(Paragraph(thank_you_text, styles['thank_you']))
                story.append(Spacer(1, 25))
                
                # Section header
                story.append(Paragraph(f"Evaluation Questions & Responses - {domain_name}", styles['section_header']))
                story.append(Paragraph("This section presents overall information regarding the tool. These answers will give you a general idea of the tool and will support your evaluation.", styles['section_desc']))
                story.append(Spacer(1, 15))
                
                # Process questions
                code_col = 9 if maturity_key == "advanced" else 8
                question_col = 11 if maturity_key == "advanced" else 10
                answer_col = 12 if maturity_key == "advanced" else 11
                
                question_count = 0
                previous_content_type = None
                domain_subheading_counter = 0
                is_after_indicator = False
                
                for row_idx in domain_info['answered_rows'][3:]:
                    try:
                        # Get question code
                        question_code = None
                        code_value = ws_answers.cell(row_idx, code_col).value
                        if code_value and str(code_value).strip().isdigit():
                            if len(str(code_value).strip()) == 8:
                                question_code = str(code_value).strip()
                        
                        # Get question text
                        question_text = ws_answers.cell(row_idx, question_col).value
                        if question_text:
                            question_text = str(question_text).strip()
                        
                        # Get answer text
                        answer_text = ws_answers.cell(row_idx, answer_col).value
                        if answer_text:
                            answer_text = str(answer_text).strip()
                        
                        # Use HTML version if available
                        if question_code and question_code in html_mapping:
                            html_text = html_mapping[question_code]
                            if len(html_text) > len(question_text or ""):
                                question_text = html_text
                        
                        if question_text and len(question_text.strip()) > 3:
                            clean_question = clean_html_text(question_text)
                            clean_question = re.sub(r'\[.*?\]', '', clean_question).strip()
                            
                            if clean_question:
                                content_type = identify_content_type(clean_question, answer_text, previous_content_type)
                                
                                if content_type == "dimension":
                                    heading_text = clean_question
                                    if "dimension:" in heading_text.lower():
                                        heading_text = heading_text.split(":", 1)[1].strip()
                                    story.append(Paragraph(f"Dimension - {heading_text}", styles['main_heading']))
                                    story.append(Spacer(1, 10))
                                    is_after_indicator = False
                                
                                elif content_type == "subdimension":
                                    heading_text = clean_question
                                    if "sub-dimension:" in heading_text.lower():
                                        parts = heading_text.split(":", 1)
                                        if len(parts) > 1:
                                            heading_text = f"Sub-dimension: {parts[1].strip()}"
                                    story.append(Paragraph(heading_text, styles['sub_heading']))
                                    story.append(Spacer(1, 8))
                                    is_after_indicator = False
                                
                                elif content_type == "indicator":
                                    story.append(Paragraph(clean_question, styles['indicator']))
                                    story.append(Spacer(1, 6))
                                    is_after_indicator = True
                                
                                elif content_type == "domain_heading":
                                    story.append(Paragraph("Domain-Specific Questions", styles['domain_heading']))
                                    story.append(Spacer(1, 15))
                                    domain_subheading_counter = 0
                                    is_after_indicator = False
                                
                                elif content_type == "domain_subheading":
                                    domain_subheading_counter += 1
                                    clean_heading = re.sub(r'^\d+\.\s*', '', clean_question.strip())
                                    formatted_heading = f"{domain_subheading_counter:02d}. {clean_heading}"
                                    story.append(Paragraph(formatted_heading, styles['domain_subheading']))
                                    story.append(Spacer(1, 10))
                                    is_after_indicator = False
                                
                                elif content_type == "additional_info_heading":
                                    story.append(Paragraph("ADDITIONAL INFORMATION", styles['additional_info']))
                                    story.append(Spacer(1, 12))
                                    is_after_indicator = False
                                
                                elif content_type == "evaluator_affirmation":
                                    story.append(Paragraph("Evaluator Affirmation", styles['evaluator']))
                                    story.append(Spacer(1, 6))
                                    is_after_indicator = False
                                
                                elif content_type == "evaluator_affirmation_text":
                                    story.append(Paragraph(clean_question, styles['evaluator_text']))
                                    story.append(Spacer(1, 8))
                                    is_after_indicator = False
                                
                                elif content_type == "innovator_response":
                                    story.append(Paragraph("Innovator Response", styles['innovator']))
                                    story.append(Spacer(1, 6))
                                    is_after_indicator = False
                                
                                elif content_type == "contextual_information":
                                    story.append(Paragraph("Contextual Information:", styles['context_heading']))
                                    story.append(Spacer(1, 6))
                                    is_after_indicator = False
                                
                                elif content_type == "description":
                                    description_text = clean_question
                                    if description_text.lower().startswith("description:"):
                                        description_text = description_text[12:].strip()
                                    story.append(Paragraph(description_text, styles['description']))
                                    story.append(Spacer(1, 12))
                                    is_after_indicator = False
                                
                                else:
                                    # Regular question-answer
                                    if is_after_indicator and not answer_text:
                                        prefixed_text = f"Description: {clean_question}"
                                        story.append(Paragraph(prefixed_text, styles['description']))
                                        story.append(Spacer(1, 8))
                                        is_after_indicator = False
                                    else:
                                        if answer_text and answer_text.strip() != "":
                                            clean_answer = clean_html_text(answer_text)
                                            if clean_answer and clean_answer != "--":
                                                answer_formatted = f'<font color="#059669"><i>"{clean_answer}"</i></font>'
                                            else:
                                                answer_formatted = '<font color="#dc2626"><i>"No response provided"</i></font>'
                                        else:
                                            answer_formatted = '<font color="#dc2626"><i>"No response provided"</i></font>'
                                        
                                        question_number = f'<font>{question_count + 1} - </font>'
                                        formatted_question = f'<font>{clean_question}</font>'
                                        combined_line = f'{question_number} {formatted_question} — {answer_formatted}'
                                        
                                        story.append(Paragraph(combined_line, styles['qa']))
                                        story.append(Spacer(1, 6))
                                        question_count += 1
                                        is_after_indicator = False
                                
                                previous_content_type = content_type
                    
                    except Exception as e:
                        debug(f"Error processing row {row_idx}: {e}")
                        continue
                
                # Footer
                story.append(Spacer(1, 30))
                footer_info = f"""
                <i><font color='#6b7280' size='9'>
                This report contains structured evaluation content from the {domain_name} domain expert.<br/>
                Generated automatically from MDII assessment data on {time.strftime('%Y-%m-%d %H:%M:%S')}.
                </font></i>
                """
                story.append(Paragraph(footer_info, styles['footer']))
                
                # Build PDF
                doc.build(story, canvasmaker=NumberedCanvas)
                debug(f"Generated PDF: {pdf_path}")
                pdf_count += 1
            
            except Exception as e:
                debug(f"Error creating PDF for domain {domain_name}: {e}")
                continue
        
        wb.close()
        debug(f"Created {pdf_count} domain-specific PDFs in: {pdf_dir}")
        return pdf_count > 0
    
    except Exception as e:
        debug(f"Error processing Excel for PDFs: {e}")
        return False

if __name__ == "__main__":
    # Test code
    import sys
    if len(sys.argv) >= 4:
        tool_id = sys.argv[1]
        excel_path = Path(sys.argv[2])
        maturity = sys.argv[3]
        generate_pdfs_from_excel(tool_id, excel_path, maturity)